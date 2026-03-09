from __future__ import annotations

import csv
import io
import json
import re
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from src.core.db import get_session
from src.core.errors import ValidationError
from src.core.file_limits import ensure_import_limits, get_extension
from src.models.strategy import Strategy
from src.services.audit_service import write_audit_log


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="ignore")


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = _decode_text(content)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValidationError("missing csv headers")
    return [row for row in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("xlsx not supported") from exc
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValidationError("empty xlsx")
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data = {}
        for idx, value in enumerate(row):
            header = headers[idx] if idx < len(headers) else ""
            if not header:
                continue
            data[header] = value
        if data:
            data_rows.append(data)
    return data_rows


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    logic_desc = (
        row.get("logic_desc")
        or row.get("策略描述")
        or row.get("策略逻辑")
        or row.get("logic")
        or row.get("strategy_desc")
        or row.get("description")
        or row.get("策略内容")
    )
    logic_code = row.get("logic_code") or row.get("strategy_code") or row.get("代码模板") or row.get("code")
    return {
        "name": row.get("name") or row.get("strategy_name") or row.get("策略名称"),
        "logic_desc": logic_desc,
        "logic_code": logic_code,
        "params_json": row.get("params_json") or row.get("params") or row.get("参数"),
        "status": row.get("status") or row.get("状态"),
    }


def _parse_json(content: bytes) -> list[dict[str, Any]]:
    text = _decode_text(content)
    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        raise ValidationError(f"invalid json: {e}") from e
    if isinstance(data, dict):
        return [data]
    if isinstance(data, list):
        if not all(isinstance(item, dict) for item in data):
            raise ValidationError("json array items must be objects")
        return data
    raise ValidationError("json must be object or array")


def _parse_params(raw: Any) -> dict[str, Any]:
    if raw is None:
        return {}
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return {}
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            return {}
    return {}


async def import_strategies(
    tenant_id: str,
    user_id: str,
    filename: str,
    content: bytes,
) -> dict[str, Any]:
    ensure_import_limits(filename, len(content), {".csv", ".xlsx", ".json"})
    ext = get_extension(filename)
    is_json = ext == ".json"
    if ext == ".csv":
        rows = _parse_csv(content)
    elif ext == ".xlsx":
        rows = _parse_xlsx(content)
    elif ext == ".json":
        rows = _parse_json(content)
    else:
        raise ValidationError("unsupported strategy file type")

    created = 0
    errors: list[str] = []

    async for session in get_session():
        created = await _insert_strategies(
            session, tenant_id, user_id, rows, errors,
            use_item_index=is_json, force_status_draft=is_json,
        )
        await session.commit()
        break

    await write_audit_log(
        tenant_id=tenant_id,
        actor_id=user_id,
        action="import_strategy",
        resource_type="strategy",
        resource_id="batch",
        status="success" if created > 0 else "failed",
        ip_address="unknown",
        user_agent="unknown",
        detail=json.dumps(
            {"imported_count": created, "error_count": len(errors)},
            ensure_ascii=False,
        ),
    )

    return {
        "imported_count": created,
        "error_count": len(errors),
        "errors": errors[:20],
    }


async def _insert_strategies(
    session: AsyncSession,
    tenant_id: str,
    user_id: str,
    rows: list[dict[str, Any]],
    errors: list[str],
    use_item_index: bool = False,
    force_status_draft: bool = False,
) -> int:
    created = 0
    start_idx = 1 if use_item_index else 2
    err_prefix = "item" if use_item_index else "row"
    for index, row in enumerate(rows, start=start_idx):
        normalized = _normalize_row(row)
        name = str(normalized.get("name") or "").strip()
        logic_desc = str(normalized.get("logic_desc") or "").strip()
        logic_code = str(normalized.get("logic_code") or "").strip()
        if logic_desc and not logic_code:
            mixed_desc, mixed_code = _split_mixed_logic(logic_desc)
            if mixed_code:
                logic_desc = mixed_desc or logic_desc
                logic_code = mixed_code
            elif not _looks_like_plain_text(logic_desc):
                logic_code = logic_desc
                logic_desc = ""
        elif logic_code and not logic_desc:
            mixed_desc, mixed_code = _split_mixed_logic(logic_code)
            if mixed_desc and mixed_code:
                logic_desc = mixed_desc
                logic_code = mixed_code
        if logic_code and not logic_desc and _looks_like_plain_text(logic_code):
            logic_desc = logic_code
            logic_code = ""
        if not name or (not logic_desc and not logic_code):
            errors.append(f"{err_prefix} {index}: missing name or strategy logic")
            continue
        params = _parse_params(normalized.get("params_json"))
        status = "draft" if force_status_draft else (
            str(normalized.get("status") or "draft").strip() or "draft"
        )
        record = Strategy(
            tenant_id=tenant_id,
            owner_id=user_id,
            name=name,
            logic_desc=logic_desc or None,
            logic_code=logic_code,
            params_json=params,
            status=status,
            is_deleted=False,
        )
        session.add(record)
        created += 1
    return created


def _looks_like_plain_text(text: str) -> bool:
    stripped = text.strip()
    if not stripped:
        return False
    code_markers = (
        "def ",
        "class ",
        "return ",
        "ctx.",
        "pass",
        "import ",
        "if ",
        "for ",
        "while ",
        "try:",
        "except",
        "lambda ",
        "=>",
        "{",
        "}",
        ";",
    )
    return not any(marker in stripped for marker in code_markers)


def _split_mixed_logic(text: str) -> tuple[str, str]:
    stripped = text.strip()
    if not stripped:
        return "", ""
    if "```" in stripped:
        parts = stripped.split("```")
        desc_parts: list[str] = []
        code_parts: list[str] = []
        for index, part in enumerate(parts):
            chunk = part.strip()
            if not chunk:
                continue
            if index % 2 == 1:
                chunk = re.sub(r"^[A-Za-z0-9_+-]+\r?\n", "", chunk, count=1)
                code_parts.append(chunk)
            else:
                desc_parts.append(chunk)
        return "\n\n".join(desc_parts).strip(), "\n\n".join(code_parts).strip()
    lines = stripped.splitlines()
    for index, line in enumerate(lines):
        if _looks_like_code_line(line):
            desc = "\n".join(lines[:index]).strip()
            code = "\n".join(lines[index:]).strip()
            if desc and code:
                return desc, code
            break
    return "", ""


def _looks_like_code_line(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    code_starts = (
        "def ",
        "class ",
        "if ",
        "elif ",
        "else:",
        "for ",
        "while ",
        "try:",
        "except",
        "with ",
        "return ",
        "import ",
        "from ",
        "pass",
    )
    return (
        stripped.startswith(code_starts)
        or "ctx." in stripped
        or bool(re.match(r"^[A-Za-z_][A-Za-z0-9_]*\s*=", stripped))
    )
